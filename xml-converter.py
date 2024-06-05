import openpyxl
import xml.etree.ElementTree as ET
import tkinter as tk
from tkinter import filedialog,messagebox,ttk
import os

# Convert the XLSX file to a custom XML format

def xlsm_to_custom_xml(xlsm_file, xml_file):
    try:
        # Read the CSV file
        workbook = openpyxl.load_workbook(xlsm_file)
        sheet = workbook.active
        # Create the root element of the XML
        deftable = ET.Element("DEFTABLE", {
            "xmlns:xsi": "http://www.w3.org/2001/XMLSchema-instance",
            "xsi:noNamespaceSchemaLocation": "Folder.xsd"
        })
        # Assuming the folder attributes are static for simplicity
        for row in sheet.iter_rows(min_row=7, values_only=True):
            if row[0] is not None:
                folder_attributes = {
                    "DATACENTER": row[5],
                    "VERSION": "920",
                    "PLATFORM": "UNIX",
                    "FOLDER_NAME": row[6],
                    "MODIFIED": "False",
                    "TYPE": "1",
                    "USED_BY_CODE": "0",
                    "ENFORCE_VALIDATION": "N"
                }
        
        folder = ET.SubElement(deftable, "FOLDER", folder_attributes)
        # Iterate over the rows of the CSV
        for row in sheet.iter_rows(min_row=7, values_only=True):
            if row[0] is not None:
                job_attributes = {
                    "JOBISN": str(row[0]),
                    "APPLICATION": row[7],
                    "SUB_APPLICATION": row[8],
                    "MEMNAME": row[10],
                    "JOBNAME": row[9],
                    "DESCRIPTION": row[17],
                    "RUN_AS": row[13],
                    "TASKTYPE": row[12],
                    "NODEID": row[15],
                    "MEMLIB": row[11],
                    "JAN": "1",
                    "FEB": "1",
                    "MAR": "1",
                    "APR": "1",
                    "MAY": "1",
                    "JUN": "1",
                    "JUL": "1",
                    "AUG": "1",
                    "SEP": "1",
                    "OCT": "1",
                    "NOV": "1",
                    "DEC": "1",
                    "PARENT_FOLDER": row[6]
                }
                job = ET.SubElement(folder, "JOB", job_attributes)
                # Example INCOND
                incond_raws=[]
                incond=row[23]
                if incond is not None:
                    incond_raw = row[23].split('\n')
                    incond_raws = incond_raw
                for i in incond_raws:
                    if i:   
                        incond_attributes = {
                            "ODATE": "ODAT",
                            "NAME": i,
                            "AND_OR": "A"
                        }
                        ET.SubElement(job, "INCOND", incond_attributes)
                # Example OUTCOND (Assuming static for simplicity)
                outcond_raws=[]
                outcond=row[25]
                if outcond == '' or None:
                    messagebox.showerror("Error", "OUTCOND is empty")
                    return 0
                outcond_raw = row[25].split('\n')
                outcond_raws = outcond_raw 
                for i in outcond_raws: 
                    if i:
                        outcond_attributes = {
                            "NAME": i,  # Assuming the OUTCOND NAME is in rowumn 28 (index 27)
                            "ODATE": "ODAT",  # Assuming the ODATE is in rowumn 29 (index 28)
                            "SIGN": "+"  # Assuming the SIGN is in rowumn 30 (index 29)
                        }
                        ET.SubElement(job, "OUTCOND", outcond_attributes)
        # Create an ElementTree object from the root element
        tree = ET.ElementTree(deftable)
        # Write the XML to the specified file
        with open(xml_file, "wb") as f:
            tree.write(f, encoding="utf-8", xml_declaration=True)
        os.startfile(xml_file)
        return f"File converted and saved to {xml_file}"
    
    except Exception as e:
        messagebox.showerror("Error", f"{e}")
        return None



#################################################################################################################################


def upload_file():
   file_path = filedialog.askopenfilename(
       filetypes=[("Excel Macro-Enabled Workbook", "*.xlsm"), ("All Files", "*.*")]
   )
   if file_path:
       selected_file_label.config(text=f"Selected File: {file_path}")
       confirm_button.config(state=tk.NORMAL)
       root.selected_file = file_path  # Store the file path in the root window object
# Function to handle the conversion after confirmation
def confirm_file():
    file_path = root.selected_file
    output_path = filedialog.asksaveasfilename(
        defaultextension=".xml",
        filetypes=[("XML Files", "*.xml"), ("All Files", "*.*")]
   )
    if output_path:
        result = xlsm_to_custom_xml(file_path, output_path)
        result_label.config(text=result)
        confirm_button.config(state=tk.DISABLED)
        

        

if __name__ == "__main__":
    root = tk.Tk()
    root.title("XLSX to XML")
    root.geometry("560x270")
    upload_button = tk.Button(root, text="Upload XLSM File", command=upload_file)
    upload_button.pack(pady=10)
    selected_file_label = tk.Label(root, text="No file selected")
    selected_file_label.pack(pady=5)
    confirm_button = tk.Button(root, text="Convert", state=tk.DISABLED, command=confirm_file)
    confirm_button.pack(pady=10)

    result_label = tk.Label(root, text="")
    result_label.pack(pady=10)

    
    root.mainloop()